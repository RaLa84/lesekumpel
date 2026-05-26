"""
Build-Script für den vollwertigen 4-Tab-Geschäftsplan Lesekumpel.

Erzeugt: geschaeftsplan-lesekumpel.xlsx mit 4 Tabs (DIN-konforme Struktur
wie INSPIRED-Wettbewerb erwartet):
  1. Personal, Investitionen etc.
  2. Mikromezzanin (statt Annuitätendarlehen)
  3. Liquiditätsplan (Cashflow-Sicht)
  4. Rentabilitätsplan (GuV-Sicht)

Alle Annahmen sind in geschaeftsplan-annahmen.md dokumentiert.
Bei Änderungen: Annahmen unten anpassen + Script neu ausführen.

Voraussetzung: pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# ============================================================
# ZENTRALE ANNAHMEN
# ============================================================

# Wachstum (avg Werte pro Jahr)
PREMIUM_AVG = {1: 50, 2: 300, 3: 1000, 4: 1700}
PRO_AVG = {1: 1, 2: 5, 3: 20, 4: 30}
KAFFEE_AVG = {1: 3, 2: 15, 3: 50, 4: 80}
FREE_RATIO = 10  # Free = 10× Premium

STORIES_PREMIUM = 10  # avg pro Monat (Cap 15)
STORIES_FREE = 3      # avg pro Monat (Cap 5)

# Preise (€/Monat bzw. €/Stück)
PREIS_PREMIUM = 4.99
PREIS_PRO = 14.99
PREIS_KAFFEE = 1.00

# Story-Kosten (mit Phase-2-Hebel ab J3)
STORY_KOSTEN = {1: 0.12, 2: 0.12, 3: 0.07, 4: 0.07}

# Investitionen (einmalig)
INVEST_MACBOOKS = 4000
INVEST_MONITORE = 1000
INVEST_MARKE = 290
INVEST_DOMAIN = 200
INVEST_TEMPLATES = 500
INVEST_SUMME = INVEST_MACBOOKS + INVEST_MONITORE + INVEST_MARKE + INVEST_DOMAIN + INVEST_TEMPLATES

# Abschreibungen (Jahre)
AFA_MACBOOKS_JAHRE = 3
AFA_MONITORE_JAHRE = 5
AFA_MARKE_JAHRE = 10

# Mikromezzanin (Phase 3, Monat 25)
MEZZ_KAPITAL = 50000
MEZZ_ZINS_PA = 0.08          # 8% p.a.
MEZZ_LAUFZEIT_MONATE = 120   # 10 Jahre
MEZZ_AUSZAHLUNG_MONAT = 25   # = Phase 3 Start
MEZZ_RUECKZAHLUNG_MONAT = MEZZ_AUSZAHLUNG_MONAT + MEZZ_LAUFZEIT_MONATE - 1  # Monat 144
MEZZ_ZINS_MONATLICH = MEZZ_KAPITAL * MEZZ_ZINS_PA / 12  # 333,33 €

# Personal (€/Monat)
GRUENDER_ENTNAHME_J3 = 3000   # 3 Gründer × 1.000 €
MITARBEITER_J3 = 3000          # ab Q3 J3 → 3 Monate

# Eigenmittel-Startpuffer
EIGENMITTEL_START = 3000

# ============================================================
# STYLES
# ============================================================

# Lesekumpel-Farben
NAVY = "2B3140"
CREAM = "FFF6EF"
MINT = "2FB8A6"
CORAL = "F97352"
YELLOW = "FFD95A"
LILA = "7D6AE6"

font_header = Font(name='Arial', size=11, bold=True, color=CREAM)
font_section = Font(name='Arial', size=10, bold=True, color=NAVY)
font_normal = Font(name='Arial', size=10, color=NAVY)
font_sum = Font(name='Arial', size=10, bold=True, color=NAVY)
font_neg = Font(name='Arial', size=10, color=CORAL, bold=True)

fill_header = PatternFill('solid', fgColor=NAVY)
fill_section = PatternFill('solid', fgColor="E8E4DC")
fill_sum = PatternFill('solid', fgColor=YELLOW)
fill_revenue = PatternFill('solid', fgColor="D7F1ED")  # helles Mint
fill_balance_pos = PatternFill('solid', fgColor="E7E2F9")  # helles Lila
fill_balance_neg = PatternFill('solid', fgColor="FCD9CC")  # helles Coral

align_right = Alignment(horizontal='right', vertical='center')
align_left = Alignment(horizontal='left', vertical='center')
align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='D9D4CC'),
    right=Side(style='thin', color='D9D4CC'),
    top=Side(style='thin', color='D9D4CC'),
    bottom=Side(style='thin', color='D9D4CC'),
)


def euro_fmt():
    return '#,##0.00 "€";[Red]-#,##0.00 "€";"–"'


def percent_fmt():
    return '0.00%'


# ============================================================
# TAB 1: Personal, Investitionen etc.
# ============================================================

def build_tab1(wb):
    ws = wb.create_sheet("Personal, Investitionen etc.")

    # Block A: Investitionen
    ws['A1'] = "Investitionen:"
    ws['A1'].font = font_section
    ws['C1'] = "Betrag"
    ws['C1'].font = font_section

    row = 2
    invest = [
        ("Rechner (2× MacBooks):", INVEST_MACBOOKS),
        ("Monitore/Peripherie:", INVEST_MONITORE),
        ("Markenanmeldung DPMA:", INVEST_MARKE),
        ("Domain/Hosting-Setup:", INVEST_DOMAIN),
        ("Rechts-Templates (eRecht24):", INVEST_TEMPLATES),
    ]
    for label, value in invest:
        ws.cell(row=row, column=2, value=label).font = font_normal
        ws.cell(row=row, column=3, value=value).number_format = euro_fmt()
        ws.cell(row=row, column=3).font = font_normal
        ws.cell(row=row, column=3).alignment = align_right
        row += 1

    ws.cell(row=row, column=2, value="Summe").font = font_sum
    ws.cell(row=row, column=3, value=INVEST_SUMME).number_format = euro_fmt()
    ws.cell(row=row, column=3).font = font_sum
    ws.cell(row=row, column=3).fill = fill_sum
    ws.cell(row=row, column=3).alignment = align_right
    ws.cell(row=row, column=5, value="Finanzierung Phase 1+2: Eigenmittel (Bootstrap). Phase 3: Mikromezzanin 50.000 € (siehe Tab 2).").font = font_normal
    row += 2

    # Block B: Kalkulatorische Abschreibungen
    ws.cell(row=row, column=1, value="Kalkulatorische Abschreibung:").font = font_section
    ws.cell(row=row, column=3, value="Rechner").font = font_section
    ws.cell(row=row, column=4, value="Monitore").font = font_section
    ws.cell(row=row, column=5, value="Marke DPMA").font = font_section
    row += 1
    ws.cell(row=row, column=2, value="Jahre:").font = font_normal
    ws.cell(row=row, column=3, value=AFA_MACBOOKS_JAHRE).font = font_normal
    ws.cell(row=row, column=4, value=AFA_MONITORE_JAHRE).font = font_normal
    ws.cell(row=row, column=5, value=AFA_MARKE_JAHRE).font = font_normal
    row += 1
    ws.cell(row=row, column=2, value="Abschreibungsbetrag je Jahr:").font = font_normal
    afa_macbooks = INVEST_MACBOOKS / AFA_MACBOOKS_JAHRE
    afa_monitore = INVEST_MONITORE / AFA_MONITORE_JAHRE
    afa_marke = INVEST_MARKE / AFA_MARKE_JAHRE
    ws.cell(row=row, column=3, value=afa_macbooks).number_format = euro_fmt()
    ws.cell(row=row, column=4, value=afa_monitore).number_format = euro_fmt()
    ws.cell(row=row, column=5, value=afa_marke).number_format = euro_fmt()
    row += 2
    ws.cell(row=row, column=2, value="Summe AfA pro Jahr:").font = font_sum
    afa_summe = afa_macbooks + afa_monitore + afa_marke
    ws.cell(row=row, column=3, value=afa_summe).number_format = euro_fmt()
    ws.cell(row=row, column=3).font = font_sum
    ws.cell(row=row, column=3).fill = fill_sum
    row += 3

    # Block C: Einnahmen
    ws.cell(row=row, column=1, value="Einnahmen").font = font_section
    ws.cell(row=row, column=3, value="Jahr 1").font = font_section
    ws.cell(row=row, column=4, value="Jahr 2").font = font_section
    ws.cell(row=row, column=5, value="Jahr 3").font = font_section
    row += 1

    einnahmen_zeilen = [
        ("Familienpaket (Premium 4,99 €)",
         12 * PREIS_PREMIUM * PREMIUM_AVG[1],
         12 * PREIS_PREMIUM * PREMIUM_AVG[2],
         12 * PREIS_PREMIUM * PREMIUM_AVG[3]),
        ("Pro-Accounts (14,99 €)",
         12 * PREIS_PRO * PRO_AVG[1],
         12 * PREIS_PRO * PRO_AVG[2],
         12 * PREIS_PRO * PRO_AVG[3]),
        ("Kaffeekassen (1 €)",
         12 * PREIS_KAFFEE * KAFFEE_AVG[1],
         12 * PREIS_KAFFEE * KAFFEE_AVG[2],
         12 * PREIS_KAFFEE * KAFFEE_AVG[3]),
        ("Förderung (Aktion Mensch/DATIpilot)", 0, 0, 5000),
    ]
    for label, j1, j2, j3 in einnahmen_zeilen:
        ws.cell(row=row, column=2, value=label).font = font_normal
        for col, val in zip([3, 4, 5], [j1, j2, j3]):
            ws.cell(row=row, column=col, value=val).number_format = euro_fmt()
            ws.cell(row=row, column=col).font = font_normal
            ws.cell(row=row, column=col).alignment = align_right
        row += 1

    ws.cell(row=row, column=2, value="Summe Einnahmen (netto)").font = font_sum
    for col, jahr in zip([3, 4, 5], [1, 2, 3]):
        formel = f"=SUM({get_column_letter(col)}{row-4}:{get_column_letter(col)}{row-1})"
        ws.cell(row=row, column=col, value=formel).number_format = euro_fmt()
        ws.cell(row=row, column=col).font = font_sum
        ws.cell(row=row, column=col).fill = fill_revenue
    row += 3

    # Block D: Ausgaben
    ws.cell(row=row, column=1, value="Ausgaben").font = font_section
    ws.cell(row=row, column=3, value="Jahr 1").font = font_section
    ws.cell(row=row, column=4, value="Jahr 2").font = font_section
    ws.cell(row=row, column=5, value="Jahr 3").font = font_section
    row += 1

    # Token-Berechnung
    def token_jahr(j):
        prem_stories = PREMIUM_AVG[j] * STORIES_PREMIUM * 12
        free_stories = PREMIUM_AVG[j] * FREE_RATIO * STORIES_FREE * 12
        return (prem_stories + free_stories) * STORY_KOSTEN[j]

    def stripe_jahr(j):
        umsatz = (12 * PREIS_PREMIUM * PREMIUM_AVG[j] +
                  12 * PREIS_PRO * PRO_AVG[j] +
                  12 * PREIS_KAFFEE * KAFFEE_AVG[j])
        transaktionen = 12 * (PREMIUM_AVG[j] + PRO_AVG[j]) + 12 * KAFFEE_AVG[j]
        return 0.014 * umsatz + 0.25 * transaktionen

    ausgaben_zeilen = [
        ("LLM-Token + Bildgenerierung", token_jahr(1), token_jahr(2), token_jahr(3)),
        ("Stripe-Gebühren (1,4 % + 0,25 €)", stripe_jahr(1), stripe_jahr(2), stripe_jahr(3)),
        ("Hosting · Backend · CDN", 25*12, 75*12, 400*12),
        ("Recht · Buchhaltung · Versicherung (GbR ab J1 · UG ab J2)", 50*12, 250*12, 500*12),
        ("Dev-Tools · Subscriptions (Claude Max etc.)", 100*12, 120*12, 150*12),
        ("Marketing · Merch · Messen", 10*12, 100*12, 500*12),
        ("Strom · Internet · Reisekosten · Sonstige", 50*12, 80*12, 170*12),
        ("Gründer:innen-Entnahme (Teilzeit J3: 3×1.000 €)", 0, 0, GRUENDER_ENTNAHME_J3 * 12),
        ("Mitarbeiter:in (ab Q3 J3 für 3 Monate)", 0, 0, MITARBEITER_J3 * 3),
        ("Zinsaufwand Mikromezzanin (ab Monat 26)", 0, 0, MEZZ_ZINS_MONATLICH * 12),
    ]
    for label, j1, j2, j3 in ausgaben_zeilen:
        ws.cell(row=row, column=2, value=label).font = font_normal
        for col, val in zip([3, 4, 5], [j1, j2, j3]):
            ws.cell(row=row, column=col, value=val).number_format = euro_fmt()
            ws.cell(row=row, column=col).font = font_normal
            ws.cell(row=row, column=col).alignment = align_right
        row += 1

    ws.cell(row=row, column=2, value="Summe Ausgaben").font = font_sum
    for col in [3, 4, 5]:
        formel = f"=SUM({get_column_letter(col)}{row-len(ausgaben_zeilen)}:{get_column_letter(col)}{row-1})"
        ws.cell(row=row, column=col, value=formel).number_format = euro_fmt()
        ws.cell(row=row, column=col).font = font_sum
        ws.cell(row=row, column=col).fill = fill_sum

    # Spalten breite
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 60


# ============================================================
# TAB 2: Mikromezzanin (statt Annuitätendarlehen)
# ============================================================

def build_tab2(wb):
    ws = wb.create_sheet("Mikromezzanin")

    # Header-Box
    ws['A1'] = "Kapital:"
    ws['A1'].font = font_section
    ws['B1'] = MEZZ_KAPITAL
    ws['B1'].number_format = euro_fmt()
    ws['B1'].font = font_sum
    ws['B1'].fill = fill_sum

    ws['A2'] = "Zinssatz p.a.:"
    ws['A2'].font = font_section
    ws['B2'] = MEZZ_ZINS_PA
    ws['B2'].number_format = percent_fmt()
    ws['B2'].font = font_sum
    ws['B2'].fill = fill_sum

    ws['D1'] = "Laufzeit Monate:"
    ws['D1'].font = font_section
    ws['E1'] = MEZZ_LAUFZEIT_MONATE
    ws['E1'].font = font_sum
    ws['E1'].fill = fill_sum

    ws['D2'] = "Monatlicher Zins:"
    ws['D2'].font = font_section
    ws['E2'] = MEZZ_ZINS_MONATLICH
    ws['E2'].number_format = euro_fmt()
    ws['E2'].font = font_sum
    ws['E2'].fill = fill_sum

    ws['G1'] = "Programm: Mikromezzaninfonds II · MBG MV"
    ws['G1'].font = font_section
    ws['G2'] = "Tilgung: endfällig in Monat 144 · keine Sicherheiten · stille Beteiligung (keine Stimmrechte)"
    ws['G2'].font = font_normal

    # Tabellen-Header
    row = 5
    headers = ["Monat", "Auszahlung", "Kapital Anfang", "Zins", "Tilgung", "Kapital Ende", "Jahres-Marker"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
    row += 1

    # Tilgungsplan
    kapital_anfang = 0
    for monat in range(1, MEZZ_LAUFZEIT_MONATE + MEZZ_AUSZAHLUNG_MONAT + 1):
        # Auszahlung
        if monat == MEZZ_AUSZAHLUNG_MONAT:
            auszahlung = MEZZ_KAPITAL
            kapital_anfang = 0
            zins = 0
            tilgung = 0
            kapital_ende = MEZZ_KAPITAL
        elif monat < MEZZ_AUSZAHLUNG_MONAT:
            # Phase 1+2: noch kein Kredit
            continue
        else:
            auszahlung = 0
            kapital_anfang = kapital_ende  # vom Vormonat
            zins = kapital_anfang * MEZZ_ZINS_PA / 12
            if monat == MEZZ_RUECKZAHLUNG_MONAT:
                tilgung = kapital_anfang  # endfällig
                kapital_ende = 0
            else:
                tilgung = 0
                kapital_ende = kapital_anfang

        is_jahres_ende = (monat - MEZZ_AUSZAHLUNG_MONAT + 1) % 12 == 0 and monat >= MEZZ_AUSZAHLUNG_MONAT
        jahres_marker = f"Jahr {(monat - MEZZ_AUSZAHLUNG_MONAT) // 12 + 1}" if is_jahres_ende else ""

        ws.cell(row=row, column=1, value=monat).font = font_normal
        ws.cell(row=row, column=2, value=auszahlung if auszahlung else None).number_format = euro_fmt()
        ws.cell(row=row, column=2).font = font_normal if not auszahlung else font_sum
        if auszahlung:
            ws.cell(row=row, column=2).fill = fill_revenue
        ws.cell(row=row, column=3, value=kapital_anfang).number_format = euro_fmt()
        ws.cell(row=row, column=3).font = font_normal
        ws.cell(row=row, column=4, value=zins if zins > 0 else None).number_format = euro_fmt()
        ws.cell(row=row, column=4).font = font_normal
        ws.cell(row=row, column=5, value=tilgung if tilgung > 0 else None).number_format = euro_fmt()
        if tilgung > 0:
            ws.cell(row=row, column=5).font = font_sum
            ws.cell(row=row, column=5).fill = fill_sum
        else:
            ws.cell(row=row, column=5).font = font_normal
        ws.cell(row=row, column=6, value=kapital_ende).number_format = euro_fmt()
        ws.cell(row=row, column=6).font = font_normal
        if jahres_marker:
            ws.cell(row=row, column=7, value=jahres_marker).font = font_sum
            ws.cell(row=row, column=7).fill = fill_sum
        row += 1

    # Spalten-Breiten
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 18


# ============================================================
# TAB 3: Liquiditätsplan
# ============================================================

def build_tab3(wb):
    ws = wb.create_sheet("Liquiditätsplan")

    # Header-Zeile: Posten | Monat 1..12 | Jahr 2 | Jahr 3 | Jahr 4-10 (kompakt)
    ws['A1'] = "Liquiditätsplan (Cashflow)"
    ws['A1'].font = font_header
    ws['A1'].fill = fill_header

    row = 2
    ws.cell(row=row, column=1, value="Posten").font = font_section

    # Monate 1-12 + Jahre 2-10
    spalten = ["Posten"]
    for m in range(1, 13):
        spalten.append(f"Monat {m}")
    for j in range(2, 11):
        spalten.append(f"Jahr {j}")
    spalten.append("Summe 10J")

    for col_idx, label in enumerate(spalten[1:], start=2):
        c = ws.cell(row=row, column=col_idx, value=label)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
    row += 1

    # Premium-Wachstum linear interpoliert pro Monat
    # Monat 0 = 0, Monat 12 = 100, Monat 24 = 500, Monat 36 = 1700
    def premium_monat(m):
        if m <= 0:
            return 0
        if m <= 12:
            return round(m * 100 / 12)
        if m <= 24:
            return round(100 + (m - 12) * 400 / 12)
        if m <= 36:
            return round(500 + (m - 24) * 1200 / 12)
        return 1700

    def pro_monat(m):
        if m <= 12:
            return round(m * 3 / 12) if m > 0 else 0
        if m <= 24:
            return round(3 + (m - 12) * 5 / 12)
        if m <= 36:
            return round(8 + (m - 24) * 22 / 12)
        return 30

    def kaffee_monat(m):
        if m <= 12:
            return round(m * 5 / 12) if m > 0 else 0
        if m <= 24:
            return round(5 + (m - 12) * 25 / 12)
        if m <= 36:
            return round(30 + (m - 24) * 70 / 12)
        return 100

    def story_kost_monat(m):
        jahr = (m - 1) // 12 + 1
        jahr = min(jahr, 4)
        return STORY_KOSTEN[jahr]

    # Werte pro Monat 1-12 sammeln
    monate_data = []
    for m in range(1, 13):
        prem = premium_monat(m)
        pro = pro_monat(m)
        kaffee = kaffee_monat(m)
        free = prem * FREE_RATIO
        story_k = story_kost_monat(m)

        einn_premium = prem * PREIS_PREMIUM
        einn_pro = pro * PREIS_PRO
        einn_kaffee = kaffee * PREIS_KAFFEE
        einn_foerd = 0
        einn_darlehen = 0

        # Token-Kosten Monat
        ausg_token = (prem * STORIES_PREMIUM + free * STORIES_FREE) * story_k
        # Stripe
        umsatz_m = einn_premium + einn_pro + einn_kaffee
        transaktionen = prem + pro + kaffee
        ausg_stripe = 0.014 * umsatz_m + 0.25 * transaktionen
        # Hosting
        ausg_hosting = 25
        # Recht (GbR-Phase: 50 €/Mon)
        ausg_recht = 50
        # Dev-Tools
        ausg_tools = 100
        # Marketing
        ausg_marketing = 10
        # Sonstige
        ausg_sonst = 50
        # Personal
        ausg_personal = 0
        # Investitionen (in Monat 1-3 verteilt)
        if m == 1:
            invest = INVEST_MACBOOKS + INVEST_MONITORE
        elif m == 2:
            invest = INVEST_MARKE + INVEST_DOMAIN
        elif m == 3:
            invest = INVEST_TEMPLATES
        else:
            invest = 0
        # Mikromezzanin: noch nicht in J1
        mezz_aus = 0
        mezz_zins = 0
        mezz_tilgung = 0

        monate_data.append({
            'einn_premium': einn_premium,
            'einn_pro': einn_pro,
            'einn_kaffee': einn_kaffee,
            'einn_foerd': einn_foerd,
            'einn_darlehen': einn_darlehen,
            'invest': invest,
            'ausg_token': ausg_token,
            'ausg_stripe': ausg_stripe,
            'ausg_hosting': ausg_hosting,
            'ausg_recht': ausg_recht,
            'ausg_tools': ausg_tools,
            'ausg_marketing': ausg_marketing,
            'ausg_sonst': ausg_sonst,
            'ausg_personal': ausg_personal,
            'mezz_zins': mezz_zins,
            'mezz_tilgung': mezz_tilgung,
        })

    # Jahressummen J2-J10 berechnen
    def jahr_data(jahr):
        d = {
            'einn_premium': 12 * PREIS_PREMIUM * PREMIUM_AVG.get(jahr, 1700),
            'einn_pro': 12 * PREIS_PRO * PRO_AVG.get(jahr, 30),
            'einn_kaffee': 12 * PREIS_KAFFEE * KAFFEE_AVG.get(jahr, 80),
            'einn_foerd': 5000 if jahr == 3 else (3000 if jahr >= 4 else 0),
            'einn_darlehen': MEZZ_KAPITAL if jahr == 3 else 0,  # Auszahlung in Monat 25
            'invest': 0,
        }
        prem_avg = PREMIUM_AVG.get(jahr, 1700)
        free_avg = prem_avg * FREE_RATIO
        story_k = STORY_KOSTEN.get(jahr, 0.07)
        d['ausg_token'] = (prem_avg * STORIES_PREMIUM + free_avg * STORIES_FREE) * 12 * story_k
        umsatz_j = d['einn_premium'] + d['einn_pro'] + d['einn_kaffee']
        transaktionen_j = 12 * (prem_avg + PRO_AVG.get(jahr, 30) + KAFFEE_AVG.get(jahr, 80))
        d['ausg_stripe'] = 0.014 * umsatz_j + 0.25 * transaktionen_j
        # Skalierende Fixkosten
        hosting = {2: 75, 3: 400}.get(jahr, 600)
        d['ausg_hosting'] = hosting * 12
        recht = {2: 250, 3: 500}.get(jahr, 600)
        d['ausg_recht'] = recht * 12
        tools = {2: 120, 3: 150}.get(jahr, 200)
        d['ausg_tools'] = tools * 12
        marketing = {2: 100, 3: 500}.get(jahr, 600)
        d['ausg_marketing'] = marketing * 12
        sonst = {2: 80, 3: 170}.get(jahr, 200)
        d['ausg_sonst'] = sonst * 12
        # Personal
        if jahr == 3:
            d['ausg_personal'] = GRUENDER_ENTNAHME_J3 * 12 + MITARBEITER_J3 * 3
        elif jahr >= 4:
            d['ausg_personal'] = 4500 * 12 + MITARBEITER_J3 * 12  # leichte Steigerung
        else:
            d['ausg_personal'] = 0
        # Mikromezzanin
        if jahr == 3:
            # Auszahlung Monat 25 = Jahr 3 Monat 1 → 11 Monate Zinsen (Monat 26-36)
            d['mezz_zins'] = MEZZ_ZINS_MONATLICH * 11
        elif 4 <= jahr <= 12:
            d['mezz_zins'] = MEZZ_ZINS_MONATLICH * 12
        else:
            d['mezz_zins'] = 0
        if jahr == 12:  # Monat 144 = Jahr 12
            d['mezz_tilgung'] = MEZZ_KAPITAL
        else:
            d['mezz_tilgung'] = 0
        return d

    # Posten-Liste mit Daten-Keys
    einnahme_posten = [
        ("Premium-Abos", 'einn_premium'),
        ("Pro-Accounts", 'einn_pro'),
        ("Kaffeekassen", 'einn_kaffee'),
        ("Förderung", 'einn_foerd'),
        ("Mikromezzanin-Auszahlung", 'einn_darlehen'),
    ]
    ausgabe_posten = [
        ("Investitionen", 'invest'),
        ("LLM-Token + Bilder", 'ausg_token'),
        ("Stripe-Gebühren", 'ausg_stripe'),
        ("Hosting · Backend", 'ausg_hosting'),
        ("Recht · Buchhaltung · Versicherung", 'ausg_recht'),
        ("Dev-Tools · Subscriptions", 'ausg_tools'),
        ("Marketing · Merch", 'ausg_marketing'),
        ("Sonstige (Strom, Reise, Porto)", 'ausg_sonst'),
        ("Personal (Entnahme + Mitarbeiter)", 'ausg_personal'),
        ("Mikromezzanin-Zinsen", 'mezz_zins'),
        ("Mikromezzanin-Tilgung (endfällig)", 'mezz_tilgung'),
    ]

    def write_zeile(row, label, get_value, total_cols=22, is_einnahme=False, is_sum_row=False):
        # Spalte 1 = Label, 2-13 = Monate, 14-22 = Jahr 2-10
        c = ws.cell(row=row, column=1, value=label)
        c.font = font_sum if is_sum_row else font_normal
        if is_sum_row:
            c.fill = fill_sum if not is_einnahme else fill_revenue
        col = 2
        # Monate 1-12
        for m_idx, m_data in enumerate(monate_data):
            val = get_value(m_data, m=m_idx + 1)
            c = ws.cell(row=row, column=col, value=val if val else None)
            c.number_format = euro_fmt()
            c.font = font_sum if is_sum_row else font_normal
            if is_sum_row:
                c.fill = fill_sum if not is_einnahme else fill_revenue
            c.alignment = align_right
            col += 1
        # Jahre 2-10
        for j in range(2, 11):
            j_data = jahr_data(j)
            val = get_value(j_data, m=0)
            c = ws.cell(row=row, column=col, value=val if val else None)
            c.number_format = euro_fmt()
            c.font = font_sum if is_sum_row else font_normal
            if is_sum_row:
                c.fill = fill_sum if not is_einnahme else fill_revenue
            c.alignment = align_right
            col += 1
        # Spalte "Summe 10J"
        c = ws.cell(row=row, column=col, value=None)  # wird via Formel gesetzt
        formel_start = get_column_letter(2)
        formel_end = get_column_letter(col - 1)
        ws.cell(row=row, column=col, value=f"=SUM({formel_start}{row}:{formel_end}{row})").number_format = euro_fmt()
        ws.cell(row=row, column=col).font = font_sum
        if is_sum_row:
            ws.cell(row=row, column=col).fill = fill_sum if not is_einnahme else fill_revenue

    # Einnahmen schreiben
    ws.cell(row=row, column=1, value="EINNAHMEN").font = font_section
    ws.cell(row=row, column=1).fill = fill_section
    row += 1
    einn_start_row = row
    for label, key in einnahme_posten:
        write_zeile(row, label, lambda d, m, k=key: d.get(k, 0))
        row += 1
    einn_end_row = row - 1
    # Summe Einnahmen
    write_zeile(row, "Summe Einnahmen",
                lambda d, m, keys=[k for _, k in einnahme_posten]: sum(d.get(k, 0) for k in keys),
                is_einnahme=True, is_sum_row=True)
    summe_einn_row = row
    row += 2

    # Ausgaben schreiben
    ws.cell(row=row, column=1, value="AUSGABEN").font = font_section
    ws.cell(row=row, column=1).fill = fill_section
    row += 1
    ausg_start_row = row
    for label, key in ausgabe_posten:
        write_zeile(row, label, lambda d, m, k=key: d.get(k, 0))
        row += 1
    ausg_end_row = row - 1
    # Summe Ausgaben
    write_zeile(row, "Summe Ausgaben",
                lambda d, m, keys=[k for _, k in ausgabe_posten]: sum(d.get(k, 0) for k in keys),
                is_sum_row=True)
    summe_ausg_row = row
    row += 2

    # Liquidität im laufenden Monat/Jahr (Einnahmen - Ausgaben)
    ws.cell(row=row, column=1, value="Liquidität im laufenden Monat/Jahr").font = font_sum
    ws.cell(row=row, column=1).fill = fill_balance_pos
    col = 2
    for col_idx in range(2, 24):  # Spalte 2-23 (Monate + Jahre)
        col_letter = get_column_letter(col_idx)
        formel = f"={col_letter}{summe_einn_row}-{col_letter}{summe_ausg_row}"
        c = ws.cell(row=row, column=col_idx, value=formel)
        c.number_format = euro_fmt()
        c.font = font_sum
        c.fill = fill_balance_pos
    liq_row = row
    row += 1

    # Kumulierte Liquidität (mit Startpuffer)
    ws.cell(row=row, column=1, value=f"Kumulierte Liquidität (inkl. {EIGENMITTEL_START} € Startkapital)").font = font_sum
    ws.cell(row=row, column=1).fill = fill_balance_pos
    # Erste Spalte: Startpuffer + Liquidität Monat 1
    ws.cell(row=row, column=2, value=f"={EIGENMITTEL_START}+B{liq_row}").number_format = euro_fmt()
    ws.cell(row=row, column=2).font = font_sum
    ws.cell(row=row, column=2).fill = fill_balance_pos
    for col_idx in range(3, 24):
        col_letter = get_column_letter(col_idx)
        prev_letter = get_column_letter(col_idx - 1)
        formel = f"={prev_letter}{row}+{col_letter}{liq_row}"
        c = ws.cell(row=row, column=col_idx, value=formel)
        c.number_format = euro_fmt()
        c.font = font_sum
        c.fill = fill_balance_pos

    # Spalten-Breiten
    ws.column_dimensions['A'].width = 40
    for col_idx in range(2, 25):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12


# ============================================================
# TAB 4: Rentabilitätsplan (GuV)
# ============================================================

def build_tab4(wb):
    ws = wb.create_sheet("Rentabilitätsplan")

    ws['A1'] = "Rentabilitätsplan (GuV — kalkulatorische Sicht)"
    ws['A1'].font = font_header
    ws['A1'].fill = fill_header

    row = 2
    ws.cell(row=row, column=1, value="Posten").font = font_section

    spalten = ["Posten"]
    for m in range(1, 13):
        spalten.append(f"Monat {m}")
    for j in range(2, 11):
        spalten.append(f"Jahr {j}")
    spalten.append("Summe 10J")

    for col_idx, label in enumerate(spalten[1:], start=2):
        c = ws.cell(row=row, column=col_idx, value=label)
        c.font = font_header
        c.fill = fill_header
        c.alignment = align_center
    row += 1

    # Helpers (gleiche Logik wie Tab 3, aber:
    # - keine Darlehen-Auszahlung als Ertrag
    # - keine Tilgung als Aufwand
    # - Abschreibungen als Aufwand statt Investitionen
    def premium_monat(m):
        if m <= 0:
            return 0
        if m <= 12:
            return round(m * 100 / 12)
        if m <= 24:
            return round(100 + (m - 12) * 400 / 12)
        if m <= 36:
            return round(500 + (m - 24) * 1200 / 12)
        return 1700

    def pro_monat(m):
        if m <= 12:
            return round(m * 3 / 12) if m > 0 else 0
        if m <= 24:
            return round(3 + (m - 12) * 5 / 12)
        if m <= 36:
            return round(8 + (m - 24) * 22 / 12)
        return 30

    def kaffee_monat(m):
        if m <= 12:
            return round(m * 5 / 12) if m > 0 else 0
        if m <= 24:
            return round(5 + (m - 12) * 25 / 12)
        if m <= 36:
            return round(30 + (m - 24) * 70 / 12)
        return 100

    def story_kost_monat(m):
        jahr = (m - 1) // 12 + 1
        jahr = min(jahr, 4)
        return STORY_KOSTEN[jahr]

    # Monatliche AfA (kalkulatorisch)
    afa_monat = (INVEST_MACBOOKS / AFA_MACBOOKS_JAHRE +
                 INVEST_MONITORE / AFA_MONITORE_JAHRE +
                 INVEST_MARKE / AFA_MARKE_JAHRE) / 12

    # Monatliche Daten
    monate_data = []
    for m in range(1, 13):
        prem = premium_monat(m)
        pro = pro_monat(m)
        kaffee = kaffee_monat(m)
        free = prem * FREE_RATIO
        story_k = story_kost_monat(m)

        d = {
            'ertrag_premium': prem * PREIS_PREMIUM,
            'ertrag_pro': pro * PREIS_PRO,
            'ertrag_kaffee': kaffee * PREIS_KAFFEE,
            'ertrag_foerd': 0,
            'aufw_token': (prem * STORIES_PREMIUM + free * STORIES_FREE) * story_k,
            'aufw_hosting': 25,
            'aufw_recht': 20,
            'aufw_tools': 100,
            'aufw_marketing': 10,
            'aufw_sonst': 50,
            'aufw_personal': 0,
            'aufw_afa': afa_monat,
            'aufw_zins': 0,
        }
        umsatz_m = d['ertrag_premium'] + d['ertrag_pro'] + d['ertrag_kaffee']
        transaktionen = prem + pro + kaffee
        d['aufw_stripe'] = 0.014 * umsatz_m + 0.25 * transaktionen
        monate_data.append(d)

    def jahr_data(jahr):
        prem_avg = PREMIUM_AVG.get(jahr, 1700)
        pro_avg = PRO_AVG.get(jahr, 30)
        kaffee_avg = KAFFEE_AVG.get(jahr, 80)
        free_avg = prem_avg * FREE_RATIO
        story_k = STORY_KOSTEN.get(jahr, 0.07)

        d = {
            'ertrag_premium': 12 * PREIS_PREMIUM * prem_avg,
            'ertrag_pro': 12 * PREIS_PRO * pro_avg,
            'ertrag_kaffee': 12 * PREIS_KAFFEE * kaffee_avg,
            'ertrag_foerd': 5000 if jahr == 3 else (3000 if jahr >= 4 else 0),
        }
        umsatz_j = d['ertrag_premium'] + d['ertrag_pro'] + d['ertrag_kaffee']
        transaktionen_j = 12 * (prem_avg + pro_avg + kaffee_avg)
        d['aufw_token'] = (prem_avg * STORIES_PREMIUM + free_avg * STORIES_FREE) * 12 * story_k
        d['aufw_stripe'] = 0.014 * umsatz_j + 0.25 * transaktionen_j
        hosting = {2: 75, 3: 400}.get(jahr, 600)
        d['aufw_hosting'] = hosting * 12
        recht = {2: 250, 3: 500}.get(jahr, 600)
        d['aufw_recht'] = recht * 12
        tools = {2: 120, 3: 150}.get(jahr, 200)
        d['aufw_tools'] = tools * 12
        marketing = {2: 100, 3: 500}.get(jahr, 600)
        d['aufw_marketing'] = marketing * 12
        sonst = {2: 80, 3: 170}.get(jahr, 200)
        d['aufw_sonst'] = sonst * 12
        if jahr == 3:
            d['aufw_personal'] = GRUENDER_ENTNAHME_J3 * 12 + MITARBEITER_J3 * 3
        elif jahr >= 4:
            d['aufw_personal'] = 4500 * 12 + MITARBEITER_J3 * 12
        else:
            d['aufw_personal'] = 0
        # AfA: bis Jahr 3 voll, ab Jahr 4 ohne MacBooks (sind dann abgeschrieben)
        if jahr <= 3:
            d['aufw_afa'] = afa_monat * 12
        elif jahr <= 5:
            d['aufw_afa'] = (INVEST_MONITORE / AFA_MONITORE_JAHRE + INVEST_MARKE / AFA_MARKE_JAHRE)
        elif jahr <= 10:
            d['aufw_afa'] = INVEST_MARKE / AFA_MARKE_JAHRE
        else:
            d['aufw_afa'] = 0
        # Zinsen Mikromezzanin
        if jahr == 3:
            d['aufw_zins'] = MEZZ_ZINS_MONATLICH * 11
        elif 4 <= jahr <= 12:
            d['aufw_zins'] = MEZZ_ZINS_MONATLICH * 12
        else:
            d['aufw_zins'] = 0
        return d

    ertrag_posten = [
        ("Premium-Abos", 'ertrag_premium'),
        ("Pro-Accounts", 'ertrag_pro'),
        ("Kaffeekassen", 'ertrag_kaffee'),
        ("Förderung", 'ertrag_foerd'),
    ]
    aufwand_posten = [
        ("LLM-Token + Bilder", 'aufw_token'),
        ("Stripe-Gebühren", 'aufw_stripe'),
        ("Hosting · Backend", 'aufw_hosting'),
        ("Recht · Buchhaltung · Versicherung", 'aufw_recht'),
        ("Dev-Tools · Subscriptions", 'aufw_tools'),
        ("Marketing · Merch", 'aufw_marketing'),
        ("Sonstige", 'aufw_sonst'),
        ("Personal (Entnahme + Mitarbeiter)", 'aufw_personal'),
        ("Abschreibungen (AfA kalkulatorisch)", 'aufw_afa'),
        ("Mikromezzanin-Zinsen", 'aufw_zins'),
    ]

    def write_zeile(row, label, get_value, is_ertrag=False, is_sum_row=False):
        c = ws.cell(row=row, column=1, value=label)
        c.font = font_sum if is_sum_row else font_normal
        if is_sum_row:
            c.fill = fill_sum if not is_ertrag else fill_revenue
        col = 2
        for m_idx, m_data in enumerate(monate_data):
            val = get_value(m_data, m=m_idx + 1)
            c = ws.cell(row=row, column=col, value=val if val else None)
            c.number_format = euro_fmt()
            c.font = font_sum if is_sum_row else font_normal
            if is_sum_row:
                c.fill = fill_sum if not is_ertrag else fill_revenue
            c.alignment = align_right
            col += 1
        for j in range(2, 11):
            j_data = jahr_data(j)
            val = get_value(j_data, m=0)
            c = ws.cell(row=row, column=col, value=val if val else None)
            c.number_format = euro_fmt()
            c.font = font_sum if is_sum_row else font_normal
            if is_sum_row:
                c.fill = fill_sum if not is_ertrag else fill_revenue
            c.alignment = align_right
            col += 1
        formel_start = get_column_letter(2)
        formel_end = get_column_letter(col - 1)
        ws.cell(row=row, column=col, value=f"=SUM({formel_start}{row}:{formel_end}{row})").number_format = euro_fmt()
        ws.cell(row=row, column=col).font = font_sum
        if is_sum_row:
            ws.cell(row=row, column=col).fill = fill_sum if not is_ertrag else fill_revenue

    # ERTRÄGE
    ws.cell(row=row, column=1, value="ERTRÄGE").font = font_section
    ws.cell(row=row, column=1).fill = fill_section
    row += 1
    for label, key in ertrag_posten:
        write_zeile(row, label, lambda d, m, k=key: d.get(k, 0))
        row += 1
    write_zeile(row, "Summe Erträge",
                lambda d, m, keys=[k for _, k in ertrag_posten]: sum(d.get(k, 0) for k in keys),
                is_ertrag=True, is_sum_row=True)
    summe_ertrag_row = row
    row += 2

    # AUFWENDUNGEN
    ws.cell(row=row, column=1, value="AUFWENDUNGEN").font = font_section
    ws.cell(row=row, column=1).fill = fill_section
    row += 1
    for label, key in aufwand_posten:
        write_zeile(row, label, lambda d, m, k=key: d.get(k, 0))
        row += 1
    write_zeile(row, "Summe Aufwendungen",
                lambda d, m, keys=[k for _, k in aufwand_posten]: sum(d.get(k, 0) for k in keys),
                is_sum_row=True)
    summe_aufw_row = row
    row += 2

    # Rentabilität im laufenden Monat/Jahr
    ws.cell(row=row, column=1, value="Rentabilität im laufenden Monat/Jahr").font = font_sum
    ws.cell(row=row, column=1).fill = fill_balance_pos
    for col_idx in range(2, 24):
        col_letter = get_column_letter(col_idx)
        formel = f"={col_letter}{summe_ertrag_row}-{col_letter}{summe_aufw_row}"
        c = ws.cell(row=row, column=col_idx, value=formel)
        c.number_format = euro_fmt()
        c.font = font_sum
        c.fill = fill_balance_pos
    rent_row = row
    row += 1

    # Kumulierte Rentabilität
    ws.cell(row=row, column=1, value="Kumulierte Rentabilität").font = font_sum
    ws.cell(row=row, column=1).fill = fill_balance_pos
    ws.cell(row=row, column=2, value=f"=B{rent_row}").number_format = euro_fmt()
    ws.cell(row=row, column=2).font = font_sum
    ws.cell(row=row, column=2).fill = fill_balance_pos
    for col_idx in range(3, 24):
        col_letter = get_column_letter(col_idx)
        prev_letter = get_column_letter(col_idx - 1)
        formel = f"={prev_letter}{row}+{col_letter}{rent_row}"
        c = ws.cell(row=row, column=col_idx, value=formel)
        c.number_format = euro_fmt()
        c.font = font_sum
        c.fill = fill_balance_pos

    ws.column_dimensions['A'].width = 40
    for col_idx in range(2, 25):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12


# ============================================================
# Übersicht-Tab (zusätzlich, als erstes Sheet)
# ============================================================

def build_uebersicht(wb):
    ws = wb.active
    ws.title = "Übersicht"

    ws['A1'] = "Lesekumpel — Geschäftsplan"
    ws['A1'].font = Font(name='Arial', size=18, bold=True, color=NAVY)

    ws['A2'] = "INSPIRED Ideenwettbewerb MV 2026"
    ws['A2'].font = Font(name='Arial', size=11, italic=True, color=LILA)

    row = 4
    info = [
        ("Geschäftsmodell:", "Neuroinklusive Lese-Bibliothek für Kinder 5-10 (LRS · ADHS · Autismus)"),
        ("Pricing:", "Familienpaket 4,99 €/Mon · Pro-Account 14,99 €/Mon · Kaffeekasse 1 €/2 Credits"),
        ("Phasen:", "Konzeption (GbR · Eigenmittel) → Validierung (UG-Umfirmierung) → Skalierung (Mikromezzanin)"),
        ("Finanzierung Phase 3:", "Mikromezzaninfonds II · 50.000 € · 8% p.a. · 10 J endfällig (MBG MV)"),
        ("Break-Even:", "~200 Premium (moderate Ratio: 1 Pro + 5 Kaffee pro 100 Premium → 2 Pro + 10 Kaffee/Mon.)"),
        ("Tragfähig (Skalierung):", "1.700 Premium · 8.500 €/Monat Umsatz = 8.500 € Kosten"),
        ("", ""),
        ("Sheet-Struktur:", ""),
        ("Tab 1:", "Personal, Investitionen etc. — Übersicht 3-Jahres-Plan mit AfA"),
        ("Tab 2:", "Mikromezzanin — Tilgungsplan über 120 Monate"),
        ("Tab 3:", "Liquiditätsplan — monatlich J1, dann J2-J10 (Cashflow)"),
        ("Tab 4:", "Rentabilitätsplan — monatlich J1, dann J2-J10 (GuV mit AfA + Zinsen)"),
        ("", ""),
        ("Annahmen:", "siehe geschaeftsplan-annahmen.md"),
        ("Reproduzierbar:", "python build-geschaeftsplan.py"),
    ]
    for label, val in info:
        if label:
            ws.cell(row=row, column=1, value=label).font = font_section
            ws.cell(row=row, column=3, value=val).font = font_normal
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 90


# ============================================================
# MAIN
# ============================================================

def main():
    wb = Workbook()
    build_uebersicht(wb)
    build_tab1(wb)
    build_tab2(wb)
    build_tab3(wb)
    build_tab4(wb)

    output_path = Path(__file__).parent / "geschaeftsplan-lesekumpel.xlsx"
    wb.save(output_path)
    print(f"OK: {output_path}")
    print(f"5 Tabs: Übersicht · Personal/Invest · Mikromezzanin · Liquidität · Rentabilität")


if __name__ == "__main__":
    main()
